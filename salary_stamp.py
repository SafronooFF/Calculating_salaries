from tkinter import *
from datetime import datetime
from openpyxl import Workbook,load_workbook

window = Tk()
window.geometry('285x150')

# Получение данных времени
dt = datetime.now()

#Создание файла excel
wb = load_workbook('work.xlsx')
sheet = wb.active
sheet.title = "Данные о зарплате"
sheet.column_dimensions["A"].width = 13 #Размер столбца A

#Работа с файлом Payout
wb2 = load_workbook('payout.xlsx')
sheet2 = wb2.active
sheet2.title = "Данные о Выплатах"
sheet2.column_dimensions["A"].width = 13 #Размер столбца A

# Счетчик данных
k = sheet["D1"].value #Считает деньги и строки в excel чтобы всё шло по порядку
k2 = sheet2["C1"].value #Счёт строк в excel payout чтобы всё шло по порядку

#Данные entry
data = sheet["D2"].value

# Сумма денег
summa = k * 1500 - data

#Добавление нового дня работы и зарплаты
def new_day():
    global k
    global summa
    k += 1
    summa = k * 1500 - data #Обновление данных суммы
    sheet['D1'] = k
    sheet[f'A{k}'] = dt.strftime('%Y, %B %d')
    sheet[f'B{k}'] = 1500
    sheet["C2"] = "=SUM(B:B)-D2"
    wb.save('work.xlsx')
    wb.close()
    print('Данные успешно добавлены')

    label2.config(text = f'{summa} рублей') #Обновление данных о вылате

#Очистка данных из work и добавление данных о выплате в payout
def clean():
    global k
    global k2
    global summa
    global data
    # payout
    k2 += 1
    sheet2[f'A{k2}'] = dt.strftime('%Y, %B %d')
    sheet2[f'B{k2}'] = summa
    sheet2['C1'] = k2
    wb2.save('payout.xlsx')
    wb2.close()
    print('Данные о выплатах обновлены')
    #work
    k = 0
    data = 0
    sheet['D2'] = data
    sheet['D1'] = k
    for i in range(1, 32):
        sheet[f'A{i}'] = ''
        sheet[f'B{i}'] = ''
    wb.save('work.xlsx')
    wb.close()
    print('Таблица пуста. Вы получили выплату')

    summa = k * 1500 - data
    label2.config(text = f'{summa} рублей')

#Добавление другой выплаты
def another_payout():
    global k
    global k2
    global summa
    global data
    data = int(entry.get())
    sheet["D2"]= data
    k2 += 1
    sheet2[f'A{k2}'] = dt.strftime('%Y, %B %d')
    sheet2[f'B{k2}'] = data
    sheet2['C1'] = k2
    wb2.save('payout.xlsx')
    wb2.close()
    print(f'Добавлена выплата {data} рублей')
    wb.save('work.xlsx')
    wb.close()

    summa = k * 1500 - data  # Обновление данных суммы
    label2.config(text=f'{summa} рублей')

label = Label(text='Зарплата:', font=("Arial", 12))
label.grid(row = 0, column = 0)

label2 = Label(text = f'{summa} рублей', font=("Arial", 12))
label2.grid(row = 0, column = 1)

btn = Button(text='Добавить день работы', command = new_day)
btn.grid(row = 1, column = 0, padx=10, pady=10)

btn = Button(text='Выплата', command = clean)
btn.grid(row = 1, column = 1)

#Поле ввода отдельной суммы выплаты
entry = Entry()
entry.grid(row = 2, column = 0)

btn = Button(text='Частичная выплата', command = another_payout)
btn.grid(row = 2, column = 1, pady=10)

window.mainloop()